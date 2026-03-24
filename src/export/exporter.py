"""Excel export with professional formatting.

Generates .xlsx files with 4 sheets: All Businesses, By City, By Category, Metadata.
Light, clean design with subtle colors, proper alignment, and auto-sized columns.
"""

import logging
import os
from collections import Counter
from datetime import UTC, datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from core.searcher import Place

__all__ = ["export_excel"]

logger = logging.getLogger(__name__)

# --- Color Palette (soft, professional) ---
_BORDER_COLOR = "D6DCE4"
_HEADER_BG = "4472C4"  # Soft blue
_HEADER_FONT_COLOR = "FFFFFF"
_SUBHEADER_BG = "D9E2F3"  # Very light blue
_ALT_ROW_BG = "F2F7FB"  # Barely-there blue tint
_TOTAL_ROW_BG = "E2EFDA"  # Light green for totals
_META_LABEL_BG = "F2F2F2"  # Light gray

# --- Reusable Styles ---
_thin_border = Border(
    left=Side(style="thin", color=_BORDER_COLOR),
    right=Side(style="thin", color=_BORDER_COLOR),
    top=Side(style="thin", color=_BORDER_COLOR),
    bottom=Side(style="thin", color=_BORDER_COLOR),
)

_header_fill = PatternFill(start_color=_HEADER_BG, end_color=_HEADER_BG, fill_type="solid")
_header_font = Font(name="Calibri", bold=True, color=_HEADER_FONT_COLOR, size=11)
_header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

_subheader_fill = PatternFill(start_color=_SUBHEADER_BG, end_color=_SUBHEADER_BG, fill_type="solid")
_subheader_font = Font(name="Calibri", bold=True, color="2F5496", size=10)

_alt_row_fill = PatternFill(start_color=_ALT_ROW_BG, end_color=_ALT_ROW_BG, fill_type="solid")
_total_row_fill = PatternFill(start_color=_TOTAL_ROW_BG, end_color=_TOTAL_ROW_BG, fill_type="solid")
_meta_label_fill = PatternFill(start_color=_META_LABEL_BG, end_color=_META_LABEL_BG, fill_type="solid")

_data_font = Font(name="Calibri", size=10, color="333333")
_data_font_bold = Font(name="Calibri", size=10, color="333333", bold=True)
_link_font = Font(name="Calibri", size=10, color="2563EB", underline="single")
_number_font = Font(name="Calibri", size=10, color="333333")
_total_font = Font(name="Calibri", size=10, color="333333", bold=True)

_left_align = Alignment(horizontal="left", vertical="center")
_center_align = Alignment(horizontal="center", vertical="center")
_right_align = Alignment(horizontal="right", vertical="center")
_wrap_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

# Column definitions: (header, width, alignment)
COLUMNS = [
    ("Place ID", 22, _left_align),
    ("Name", 32, _left_align),
    ("Language", 10, _center_align),
    ("Category", 22, _left_align),
    ("All Types", 30, _wrap_align),
    ("Address", 45, _wrap_align),
    ("Region", 18, _left_align),
    ("City", 15, _left_align),
    ("Latitude", 12, _right_align),
    ("Longitude", 12, _right_align),
    ("Phone (Local)", 16, _left_align),
    ("Phone (Intl)", 18, _left_align),
    ("Rating", 8, _center_align),
    ("Reviews", 9, _right_align),
    ("Website", 30, _left_align),
    ("Google Maps", 30, _left_align),
    ("Status", 14, _center_align),
    ("Hours", 50, _wrap_align),
    ("Scraped Date", 20, _center_align),
]


def export_excel(
    places: list[Place],
    output_path: str,
    metadata: dict | None = None,
) -> str:
    """Export places to a professionally formatted Excel file."""
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb = Workbook()

    _write_all_businesses(wb, places)
    _write_by_city(wb, places)
    _write_by_category(wb, places)
    _write_metadata(wb, metadata or {})

    wb.save(output_path)
    logger.info("Exported %d businesses to %s", len(places), output_path)
    return output_path


def _apply_cell(cell, font=None, fill=None, alignment=None, border=None, number_format=None):
    """Apply styles to a cell."""
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border
    if number_format:
        cell.number_format = number_format


def _write_header_row(ws, headers, row=1):
    """Write a styled header row."""
    for col_idx, header_text in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header_text)
        _apply_cell(cell, font=_header_font, fill=_header_fill,
                    alignment=_header_alignment, border=_thin_border)
    ws.row_dimensions[row].height = 30


def _write_all_businesses(wb: Workbook, places: list[Place]) -> None:
    """Sheet 1: All Businesses — full data, professional styling."""
    ws = wb.active
    ws.title = "All Businesses"
    ws.sheet_properties.tabColor = "4472C4"
    now = datetime.now(UTC).strftime("%Y-%m-%d %H:%M UTC")

    # Header row
    _write_header_row(ws, [col[0] for col in COLUMNS])

    # Set column widths
    for col_idx, (_, width, _) in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Data rows
    for row_idx, place in enumerate(places, 2):
        values = [
            place.place_id, place.name, place.name_language,
            place.primary_type_display, ", ".join(place.types), place.address,
            place.region, place.city, place.latitude, place.longitude,
            place.phone_local, place.phone_intl, place.rating, place.review_count,
            place.website, place.google_maps_url, place.business_status, place.hours, now,
        ]

        row_fill = _alt_row_fill if row_idx % 2 == 0 else None

        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            col_align = COLUMNS[col_idx - 1][2]
            _apply_cell(cell, font=_data_font, fill=row_fill,
                        alignment=col_align, border=_thin_border)

        # Special formatting
        # Coordinates: 6 decimal places
        for coord_col in (9, 10):
            ws.cell(row=row_idx, column=coord_col).number_format = "0.000000"

        # Rating: 1 decimal
        rating_cell = ws.cell(row=row_idx, column=13)
        if rating_cell.value is not None:
            rating_cell.number_format = "0.0"

        # Reviews: thousands separator
        ws.cell(row=row_idx, column=14).number_format = "#,##0"

        # Phone columns: text format (prevent scientific notation)
        for phone_col in (11, 12):
            ws.cell(row=row_idx, column=phone_col).number_format = "@"

        # Website & Google Maps: hyperlinks
        for link_col in (15, 16):
            link_cell = ws.cell(row=row_idx, column=link_col)
            if link_cell.value:
                link_cell.font = _link_font if row_fill is None else Font(
                    name="Calibri", size=10, color="2563EB", underline="single"
                )
                link_cell.hyperlink = link_cell.value

        ws.row_dimensions[row_idx].height = 18

    # Freeze header and auto-filter
    ws.freeze_panes = "A2"
    if ws.max_row > 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{ws.max_row}"


def _write_by_city(wb: Workbook, places: list[Place]) -> None:
    """Sheet 2: Business count per city, broken down by category."""
    ws = wb.create_sheet("By City")
    ws.sheet_properties.tabColor = "548235"

    # Collect city -> category -> count
    city_cats: dict[str, Counter] = {}
    for place in places:
        city_cats.setdefault(place.city, Counter())[place.primary_type_display] += 1

    all_cats = sorted({cat for counts in city_cats.values() for cat in counts})
    headers = ["City", "Total"] + all_cats

    _write_header_row(ws, headers)

    # Column widths
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 10
    for col_idx in range(3, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(
            len(headers[col_idx - 1]) + 2, 10
        )

    # Data rows
    grand_total = 0
    cat_totals: Counter = Counter()
    for row_idx, (city, counts) in enumerate(
        sorted(city_cats.items(), key=lambda x: -sum(x[1].values())), 2
    ):
        row_fill = _alt_row_fill if row_idx % 2 == 0 else None
        city_total = sum(counts.values())
        grand_total += city_total

        cell = ws.cell(row=row_idx, column=1, value=city)
        _apply_cell(cell, font=_data_font_bold, fill=row_fill,
                    alignment=_left_align, border=_thin_border)

        cell = ws.cell(row=row_idx, column=2, value=city_total)
        _apply_cell(cell, font=_data_font_bold, fill=row_fill,
                    alignment=_right_align, border=_thin_border)
        cell.number_format = "#,##0"

        for col_idx, cat in enumerate(all_cats, 3):
            val = counts.get(cat, 0)
            cat_totals[cat] += val
            cell = ws.cell(row=row_idx, column=col_idx, value=val if val > 0 else "")
            _apply_cell(cell, font=_data_font, fill=row_fill,
                        alignment=_right_align, border=_thin_border)
            if val > 0:
                cell.number_format = "#,##0"

        ws.row_dimensions[row_idx].height = 18

    # Totals row
    total_row = len(city_cats) + 2
    cell = ws.cell(row=total_row, column=1, value="TOTAL")
    _apply_cell(cell, font=_total_font, fill=_total_row_fill,
                alignment=_left_align, border=_thin_border)
    cell = ws.cell(row=total_row, column=2, value=grand_total)
    _apply_cell(cell, font=_total_font, fill=_total_row_fill,
                alignment=_right_align, border=_thin_border)
    cell.number_format = "#,##0"
    for col_idx, cat in enumerate(all_cats, 3):
        val = cat_totals[cat]
        cell = ws.cell(row=total_row, column=col_idx, value=val if val > 0 else "")
        _apply_cell(cell, font=_total_font, fill=_total_row_fill,
                    alignment=_right_align, border=_thin_border)
        if val > 0:
            cell.number_format = "#,##0"
    ws.row_dimensions[total_row].height = 20

    ws.freeze_panes = "A2"


def _write_by_category(wb: Workbook, places: list[Place]) -> None:
    """Sheet 3: Count per primary type, sorted descending, with percentage."""
    ws = wb.create_sheet("By Category")
    ws.sheet_properties.tabColor = "BF8F00"

    counts = Counter(p.primary_type_display for p in places)
    total = len(places)
    headers = ["Category", "Count", "Percentage"]

    _write_header_row(ws, headers)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 14

    for row_idx, (cat, count) in enumerate(counts.most_common(), 2):
        row_fill = _alt_row_fill if row_idx % 2 == 0 else None
        pct = count / total if total > 0 else 0

        cell = ws.cell(row=row_idx, column=1, value=cat or "(uncategorized)")
        _apply_cell(cell, font=_data_font, fill=row_fill,
                    alignment=_left_align, border=_thin_border)

        cell = ws.cell(row=row_idx, column=2, value=count)
        _apply_cell(cell, font=_data_font, fill=row_fill,
                    alignment=_right_align, border=_thin_border)
        cell.number_format = "#,##0"

        cell = ws.cell(row=row_idx, column=3, value=pct)
        _apply_cell(cell, font=_data_font, fill=row_fill,
                    alignment=_right_align, border=_thin_border)
        cell.number_format = "0.0%"

        ws.row_dimensions[row_idx].height = 18

    # Totals row
    total_row = len(counts) + 2
    cell = ws.cell(row=total_row, column=1, value="TOTAL")
    _apply_cell(cell, font=_total_font, fill=_total_row_fill,
                alignment=_left_align, border=_thin_border)
    cell = ws.cell(row=total_row, column=2, value=total)
    _apply_cell(cell, font=_total_font, fill=_total_row_fill,
                alignment=_right_align, border=_thin_border)
    cell.number_format = "#,##0"
    cell = ws.cell(row=total_row, column=3, value=1.0 if total > 0 else 0)
    _apply_cell(cell, font=_total_font, fill=_total_row_fill,
                alignment=_right_align, border=_thin_border)
    cell.number_format = "0.0%"
    ws.row_dimensions[total_row].height = 20

    ws.freeze_panes = "A2"


def _write_metadata(wb: Workbook, metadata: dict) -> None:
    """Sheet 4: Run metadata — clean key-value layout."""
    ws = wb.create_sheet("Metadata")
    ws.sheet_properties.tabColor = "7B7B7B"

    headers = ["Property", "Value"]
    _write_header_row(ws, headers)
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 50

    defaults = {
        "Generated": datetime.now(UTC).strftime("%Y-%m-%d %H:%M UTC"),
        "Tool": "daleel (دليل)",
    }
    all_meta = {**defaults, **metadata}

    for row_idx, (key, value) in enumerate(all_meta.items(), 2):
        cell = ws.cell(row=row_idx, column=1, value=key)
        _apply_cell(cell, font=_data_font_bold, fill=_meta_label_fill,
                    alignment=_left_align, border=_thin_border)

        cell = ws.cell(row=row_idx, column=2, value=str(value))
        _apply_cell(cell, font=_data_font, alignment=_left_align, border=_thin_border)

        ws.row_dimensions[row_idx].height = 20

    ws.freeze_panes = "A2"
