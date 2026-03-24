"""Excel export with professional formatting.

Generates .xlsx files with 4 sheets: All Businesses, By City, By Category, Metadata.
"""

import logging
import os
from collections import Counter
from datetime import UTC, datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from searcher import Place

__all__ = ["export_excel"]

logger = logging.getLogger(__name__)

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
ALT_ROW_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
DEFAULT_FONT = Font(name="Arial", size=10)

COLUMNS = [
    "Place ID", "Name", "Name Language", "Category", "All Types", "Address",
    "Region", "City", "Latitude", "Longitude", "Phone (Local)", "Phone (Intl)",
    "Rating", "Reviews", "Website", "Google Maps", "Status", "Hours", "Scraped Date",
]


def export_excel(
    places: list[Place],
    output_path: str,
    metadata: dict | None = None,
) -> str:
    """Export places to a formatted Excel file.

    Args:
        places: List of Place objects to export.
        output_path: Full path for the output .xlsx file.
        metadata: Optional dict with run metadata (target, api_calls, cost, etc.).

    Returns:
        The output file path.
    """
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb = Workbook()

    _write_all_businesses(wb, places)
    _write_by_city(wb, places)
    _write_by_category(wb, places)
    _write_metadata(wb, metadata or {})

    wb.save(output_path)
    logger.info("Exported %d businesses to %s", len(places), output_path)
    return output_path


def _write_all_businesses(wb: Workbook, places: list[Place]) -> None:
    """Sheet 1: All Businesses with full data and formatting."""
    ws = wb.active
    ws.title = "All Businesses"
    now = datetime.now(UTC).strftime("%Y-%m-%d %H:%M UTC")

    # Header row
    for col_idx, header in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_idx, place in enumerate(places, 2):
        values = [
            place.place_id, place.name, place.name_language,
            place.primary_type_display, ", ".join(place.types), place.address,
            place.region, place.city, place.latitude, place.longitude,
            place.phone_local, place.phone_intl, place.rating, place.review_count,
            place.website, place.google_maps_url, place.business_status, place.hours, now,
        ]
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = DEFAULT_FONT
            if row_idx % 2 == 0:
                cell.fill = ALT_ROW_FILL

    # Freeze header and auto-filter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Auto-width columns
    for col_idx in range(1, len(COLUMNS) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            (len(str(ws.cell(row=r, column=col_idx).value or "")) for r in range(1, min(ws.max_row + 1, 100))),
            default=10,
        )
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)


def _write_by_city(wb: Workbook, places: list[Place]) -> None:
    """Sheet 2: Business count per city, broken down by category."""
    ws = wb.create_sheet("By City")

    # Collect city -> category -> count
    city_cats: dict[str, Counter] = {}
    for place in places:
        city_cats.setdefault(place.city, Counter())[place.primary_type_display] += 1

    # Get all categories
    all_cats = sorted({cat for counts in city_cats.values() for cat in counts})
    headers = ["City", "Total"] + all_cats

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    for row_idx, (city, counts) in enumerate(sorted(city_cats.items()), 2):
        ws.cell(row=row_idx, column=1, value=city).font = DEFAULT_FONT
        ws.cell(row=row_idx, column=2, value=sum(counts.values())).font = DEFAULT_FONT
        for col_idx, cat in enumerate(all_cats, 3):
            ws.cell(row=row_idx, column=col_idx, value=counts.get(cat, 0)).font = DEFAULT_FONT

    ws.freeze_panes = "A2"


def _write_by_category(wb: Workbook, places: list[Place]) -> None:
    """Sheet 3: Count per primary type, sorted descending."""
    ws = wb.create_sheet("By Category")
    counts = Counter(p.primary_type_display for p in places)

    headers = ["Category", "Count"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    for row_idx, (cat, count) in enumerate(counts.most_common(), 2):
        ws.cell(row=row_idx, column=1, value=cat).font = DEFAULT_FONT
        ws.cell(row=row_idx, column=2, value=count).font = DEFAULT_FONT

    ws.freeze_panes = "A2"


def _write_metadata(wb: Workbook, metadata: dict) -> None:
    """Sheet 4: Run metadata."""
    ws = wb.create_sheet("Metadata")

    headers = ["Property", "Value"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    defaults = {
        "Generated": datetime.now(UTC).strftime("%Y-%m-%d %H:%M UTC"),
        "Tool": "daleel (دليل)",
    }
    all_meta = {**defaults, **metadata}

    for row_idx, (key, value) in enumerate(all_meta.items(), 2):
        ws.cell(row=row_idx, column=1, value=key).font = DEFAULT_FONT
        ws.cell(row=row_idx, column=2, value=str(value)).font = DEFAULT_FONT

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 50
    ws.freeze_panes = "A2"
